#!/usr/bin/env python3

import os
import sys
import re
import glob
import yaml
from rich.console import Console
from rich.table import Table
from rich import print, box
from mergedeep import merge
import openpyxl
import requests
import inquirer
import argparse
import textwrap

from wismap.core import (
    load_data, list_modules, get_module_info, get_base_slots, combine,
    PINS_PER_TYPE, SLOT_NAMES,
)

# -----------------------------------------------------------------------------
# Configuration
# -----------------------------------------------------------------------------

data_folder = "./data"
definitions_file = f"{data_folder}/definitions.yml"
config_file = f"{data_folder}/config.yml"
patches_folder = f"{data_folder}/patches"
spreadsheet_url = "https://downloads.rakwireless.com/LoRa/WisBlock/Pin-Mapper/WisBlock-IO-Pin-Mapper.xlsx"
spreadsheet_file = f"{data_folder}/WisBlock-IO-Pin-Mapper.xlsx"
show_nc = False
table_format = box.SQUARE # box.SQUARE or box.MARKDOWN

# Load data via core
definitions, config = load_data(data_folder)

# -----------------------------------------------------------------------------
# Action LIST
# -----------------------------------------------------------------------------

def action_list():

    print()
    table = Table(box=table_format)
    for column in ['Module', "Type", "Description"]:
        table.add_column(column)
    for module in definitions.keys():
        table.add_row(*[module.upper(), definitions[module]['type'], definitions[module]['description']], style='bright_green')
    console = Console()
    console.print(table)
    print()

# -----------------------------------------------------------------------------
# Action INFO
# -----------------------------------------------------------------------------

def action_info(*args):

    # Get module
    if  len(args) > 0:
        module = args[0].lower()
        if module not in definitions:
            print(f"ERROR: specified module not found ({module})")
            return
    else:
        questions = [inquirer.List('module', message="Select module", choices=[(definitions[module]['description'], module) for module in definitions.keys()], carousel=True,)]
        answer = inquirer.prompt(questions)
        module = answer['module']

    info = get_module_info(definitions, config, module, show_nc)

    # Notes (we'll collect and print at the end, like the original)
    notes = list(info['notes'])
    for url in info.get('images', []):
        notes.append(f"Image: {url}")
    for url in info.get('schematics', []):
        notes.append(f"Schematic: {url}")

    print(f"Module: {module.upper()}")
    print(f"Type: {info['type']}")
    print(f"Description: {info['description']}")
    print(f"Documentation: {info['documentation']}")

    if info['type'] == 'WisSensor':
        print(f"Long: {info['double']}")

    if info['i2c_address']:
        print(f"I2C Address: {info['i2c_address']}")

    if info['mapping'] is not None:
        print(f"Mapping:")
        table = Table(box=table_format)
        for column in ["PIN", "Function"]:
            table.add_column(column)
        for row in info['mapping']:
            table.add_row(row['pin'], row['function'], style='bright_green')
        console = Console()
        console.print(table)

    if info['slots_table'] is not None:
        print(f"Slots:")
        table = Table(box=table_format)
        table.add_column("ID")
        for col in info['slots_table']['columns']:
            table.add_column(col)

        for row in info['slots_table']['rows']:
            table.add_row(*[row['pin']] + [row.get(col, '') for col in info['slots_table']['columns']], style='bright_green')

        console = Console()
        console.print(table)

    if len(notes):
        print(f"Notes:")
        for note in notes:
            print(f"- {note}")

    print()

# -----------------------------------------------------------------------------
# Action COMBINE
# -----------------------------------------------------------------------------

def action_combine(*args):

    # -------------------------------------------------------------------------
    # Gather info
    # -------------------------------------------------------------------------

    slot_module={}

    # Select base module
    if len(args) > 0:
        slot_module['BASE'] = args[0].lower()
    else:
        choices = [(definitions[module]['description'], module) for module in definitions.keys() if definitions[module]['type'] == 'WisBase']
        questions = [inquirer.List('output', message="Select Base Board", choices=choices, carousel=True)]
        slot_module['BASE'] = inquirer.prompt(questions)['output']

    # Get slot definitions (using deepcopy-safe core function for introspection)
    base_slots = get_base_slots(definitions, config, slot_module['BASE'])
    slot_names_list = list(base_slots.keys())

    # Do we have predefined configuration?
    if len(args) > 1:
        index = 1
        for slot in slot_names_list:
            if index >= len(args):
                slot_module[slot] = 'EMPTY'
            else:
                module = args[index].lower()
                if module == 'empty':
                    slot_module[slot] = 'EMPTY'
                else:
                    slot_module[slot] = module
            index+=1

    # Walk the different slots
    else:
        blocked = []
        # We need resolved slots for double info — get from base_slots
        for slot in slot_names_list:

            if slot in blocked:
                print(f"{slot} is blocked by another sensor\n")
                slot_module[slot] = 'BLOCKED'
                continue

            slot_info = base_slots[slot]

            if slot.startswith('CORE'):
                choices = [(definitions[module]['description'], module) for module in definitions.keys() if definitions[module]['type'] == 'WisCore']
                questions = [inquirer.List('output', message="Select Core Module", choices=choices, carousel=True)]
                slot_module[slot] = inquirer.prompt(questions)['output']

            if slot.startswith('SENSOR'):
                is_double = slot_info['double']
                is_double_text = "(double)" if is_double else ""
                choices = [(definitions[module]['description'], module) for module in definitions.keys() if (definitions[module]['type'] == 'WisSensor') and (is_double or not definitions[module].get('double', False))]
                choices.insert(0, ("Empty", "EMPTY"))
                questions = [inquirer.List('output', message=f"Select Sensor Module in slot {slot} {is_double_text}", choices=choices, carousel=True)]
                slot_module[slot] = inquirer.prompt(questions)['output']
                if slot_module[slot] != 'EMPTY':
                    if definitions[slot_module[slot]].get('double', False):
                        blocks = slot_info['double_blocks']
                        if blocks:
                            blocked.append(blocks)

            if slot.startswith('IO'):
                choices = [(definitions[module]['description'], module) for module in definitions.keys() if definitions[module]['type'] == 'WisIO']
                choices.insert(0, ("Empty", "EMPTY"))
                questions = [inquirer.List('output', message=f"Select IO Module in slot {slot}", choices=choices, carousel=True)]
                slot_module[slot] = inquirer.prompt(questions)['output']

            if slot.startswith('POWER'):
                choices = [(definitions[module]['description'], module) for module in definitions.keys() if definitions[module]['type'] == 'WisPower']
                questions = [inquirer.List('output', message=f"Select Power Module in slot {slot}", choices=choices, carousel=True)]
                slot_module[slot] = inquirer.prompt(questions)['output']

    # -------------------------------------------------------------------------
    # Build mapping via core
    # -------------------------------------------------------------------------

    # Build slot_assignments (without BASE)
    slot_assignments = {k: v for k, v in slot_module.items() if k != 'BASE'}
    result = combine(definitions, config, slot_module['BASE'], slot_assignments)

    # -------------------------------------------------------------------------
    # View
    # -------------------------------------------------------------------------

    columns = result['columns']
    function_slot = result['function_table']
    conflict_functions = result['conflicts']['functions']
    conflict_notes = result['conflicts']['notes']
    documentation = result['documentation']
    notes = result['notes']

    # Get core board mapping
    print()
    table = Table(box=table_format)
    for column in columns:
        table.add_column(column)
    table.add_row(*['MODULE']+[v.upper() for k, v in result['slot_module'].items()], style="bright_blue")

    for function, row in function_slot.items():
        style = "bright_yellow" if function in conflict_functions else "bright_green"
        table.add_row(*row, style=style)
    console = Console()
    console.print(table)

    if len(conflict_notes) or len(notes):
        print(f"Notes:")
        for conflict in conflict_notes:
            print(f"- {conflict}")
        for note in notes:
            print(f"- {note}")

    print(f"Documentation:")
    for line in documentation:
        print(f"- {line}")

    # Reproduce configuration
    print(f"Reproduce this configuration: python wismap.py combine {' '.join([v.lower() for k, v in result['slot_module'].items()])}")

# -----------------------------------------------------------------------------
# Action IMPORT
# -----------------------------------------------------------------------------

def import_sheet(data, sheet):

    # Get code
    module_code = sheet.title.lower()
    if not module_code in data:
        data[module_code] = {}

    # Get column offset
    column_offset = 1
    if sheet['A3'].value == 'PIN number':
        column_offset = 0

    # Get type
    module_type = "WisBase"
    if sheet.cell(row = 26, column = 2 + column_offset).value == 'BOOT0':
        module_type = "WisCore"
    elif sheet.cell(row = 43, column = 1 + column_offset).value == 40:
        module_type = "WisIO"
    elif sheet.cell(row = 2, column = 3 + column_offset).value == 'SLOT A':
        module_type = "WisSensor"
    data[module_code]['type'] = module_type

    # Get description
    key_column = 1 + column_offset
    value_column = 3 + column_offset
    rows = 40
    if module_type == "WisBase":
        module_description = module_code
    if module_type == "WisCore":
        module_description = module_code
    if module_type == "WisIO":
        module_description = sheet.cell(row = 45, column = 2 + column_offset).value
    if module_type == "WisSensor":
        module_description = sheet.cell(row = 29, column = 2 + column_offset).value
        rows = 24
    data[module_code]['description'] = module_description.strip(' "\'\t\r\n').replace("WisBlock ", "")

    # Get documentation
    module_docs = f"https://docs.rakwireless.com/product-categories/wisblock/{ sheet.title.lower() }/overview/"
    data[module_code]['documentation'] = module_docs

    # Get mapping
    mapping = {}
    for row in range(4, rows+4):
        pin = str(sheet.cell(row = row, column = key_column).value)
        if pin == "" or pin == "Description":
            break
        function = sheet.cell(row = row, column = value_column).value
        if pin.isnumeric():
            pin = int(pin)
        if function != "NC":
            mapping[pin] = function
    data[module_code]['mapping'] = mapping

    # I2C Address
    address = str(sheet.cell(row = row+3, column = 2+column_offset).value)
    matches = re.findall(r"0x[0-9a-fA-F]{2}", address)
    if len(matches):
        data[module_code]['i2c_address'] = matches[0]


def action_import(patch=True):

    skip_sheets = ["Pin Mapper", "model list", "NA IO", "NA_SENS"]

    print
    print("[bold cyan]---------------------------------------[/]")
    print("[bold cyan]Importing data from original spreadheet[/]")
    print("[bold cyan]---------------------------------------[/]")

    if not os.path.isfile(spreadsheet_file):
        print("Downloading spreadsheet")
        response = requests.get(spreadsheet_url)
        if not response.ok:
            sys.exit(1)
        with open(spreadsheet_file, mode="wb") as file:
            file.write(response.content)
    else:
        print("Using cached spreadsheet")

    # Open Pin Mapper spreadsheet
    wb = openpyxl.load_workbook(spreadsheet_file)
    print(f"Found {len(wb.sheetnames) - len(skip_sheets)} products")

    # Output data
    data = {}

    # Walk sheets
    for sheet_name in wb.sheetnames:
        if sheet_name not in skip_sheets:
            #print(f"Importing {sheet_name.upper()}...")
            sheet = wb[sheet_name]
            import_sheet(data, sheet)

    # Apply patches
    if patch:

        # Load patch files
        patches = {}
        for patch_file in sorted(glob.glob(f"{patches_folder}/*.yml")):
            with open(patch_file) as f:
                patch = yaml.load(f, Loader=yaml.loader.SafeLoader)
                if patch:
                    patches.update({k: v for k, v in patch.items() if v is not None})

        # Apply
        if len(patches.keys()):
            print("Applying patches")
            data = merge(data, patches)

    # Filter & sort mappings
    print("Filtering and sorting")
    for module in data:
        if 'mapping' in data[module]:
            data[module]['mapping'] = dict(sorted([(k, v) for k, v in data[module]['mapping'].items() if v is not None ]))
    data = dict(sorted(data.items(), key=lambda e: int(re.findall(r"\d+", e[0])[0])))

    # Resume
    print(f"Final list has {len(data.keys())} products")

    # Save
    with open(definitions_file, "w") as w:
        print("Saving definitions")
        yaml.dump(data, w, sort_keys=False)

# -----------------------------------------------------------------------------
# Action CLEAN
# -----------------------------------------------------------------------------

def action_clean():

    # Delete file
    if os.path.isfile(spreadsheet_file):
        os.remove(spreadsheet_file)

# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------

ACTIONS = {
    "list" : action_list,
    "info" : action_info,
    "combine" : action_combine,
    "import" : action_import,
    "clean" : action_clean,
}

parser = argparse.ArgumentParser(
    formatter_class=argparse.RawDescriptionHelpFormatter,
    prog='python wismap.py',
    usage='%(prog)s [-h] [-m] [-n] action [extra]',
    epilog=textwrap.dedent('''The 'info' action accepts the name of the module to show as an extra argument.\nThe 'combine' action accepts a list of modules to mount on the different slots, starting with the base module.''')
)
parser.add_argument('action', default='list', nargs='?', help='Action to run: '+ ', '.join(ACTIONS.keys()))
parser.add_argument('-m', '--markdown', default=False, help='Show tables in markdown format', action='store_true')
parser.add_argument('-n', '--nc', default=False, help='Show NC pins', action='store_true')
(arguments, extra) = parser.parse_known_args()

action = arguments.action
if action not in ACTIONS.keys():
    print(f"ERROR: unknown action '{action}'")
    parser.print_help()
    sys.exit(1)
show_nc = arguments.nc
table_format = box.MARKDOWN if arguments.markdown else box.SQUARE

# Execute
ACTIONS[action](*extra)
